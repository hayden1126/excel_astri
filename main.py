import openpyxl, webbrowser
from datetime import date, datetime
from dateutil.relativedelta import relativedelta

# Converts Excel columns to integer (e.g. A -> 0, B -> 1, AA -> 26, AB -> 27)
def alpha2col(col):
    col = col.upper()
    if len(col) == 1:
        return ord(col) - ord('A')
    else:
        return (ord(col[0]) - ord('A') + 1) * 26 + (ord(col[1]) - ord('A'))
int2month = {1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr", 5: "May", 6: "Jun", 7: "Jul", 8: "Aug", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec"}

optfile = "Copy of ProjectsListOnly 2023.05 (Circulated) (003).xlsx"
contractnum_column = alpha2col("F")
url = "http://k2.astri.org/FCP/E-Form/Form_Clearance_View.aspx?WF_ID="

outfile = "Copy of Copy of TT Listing_FY2019-20_Payment  deliverables schedule-r1.xlsx"
startdatecolumn = alpha2col("AU")
enddatecolumn = alpha2col("AV")

project_code_column = alpha2col("AQ")

wb_commencedate = openpyxl.load_workbook(optfile)
wb_out = openpyxl.load_workbook(outfile)

# accessing main sheet
sheet_commencedate = wb_commencedate.active # find project_code of row matching column C, and get that row
sheet_out = wb_out.active

# For each row in the sheet, open the url in a browser
for row in sheet_out.iter_rows(min_row=4): # Start at 4th row
    
    print("\n", row[0].value, row[1].value)
    
    project_code = row[project_code_column].value
    found = False
    
    if project_code != None:
        print("Project Code:", project_code)
        
        # Find row where column C matches project_code in optfile
        for row2 in sheet_commencedate.iter_rows(min_row=4): # Start at 4th row
            if row2[2].value == project_code:
                print("Commencement date:", row2[alpha2col("J")].value)
                found = True
                break

    if not found:
        print("Project code not found")
        
        contractnum = row[contractnum_column].value
        print("Launching: " + url + str(contractnum))
        webbrowser.open(url + str(contractnum))
        
        input("Press Enter to continue (after inputting commencement date to excel file)...")
        
    startdate = row[startdatecolumn].value
    print(startdate)
    years = int(input("Enter number of years: "))
    months = int(input("Enter number of months: "))
    # days = int(input("Enter number of days: "))
    
    # convert startdate in the form of 03 Jun 2023 to integer to fit date(YYYY, MM, DD)
    date_object = datetime.strptime(startdate, '%d %b %Y').date()
    enddate = date_object + relativedelta(years=years, months=months) 
    print("Enddate:", f"{enddate.day} {int2month[enddate.month]} {enddate.year}")
    
    input("Press Enter to continue (after inputting delivery due date to excel file)...")
